from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import ICPCRegistration


@admin.register(ICPCRegistration)
class ICPCRegistrationAdmin(admin.ModelAdmin):
    list_display = [
        'full_name',
        'student_id',
        'academic_class',
        'cohort',
        'phone',
        'email',
        'programming_experience',
        'status',
        'created_at',
    ]
    list_filter = ['status', 'cohort', 'programming_experience', 'created_at']
    search_fields = ['full_name', 'student_id', 'academic_class', 'email', 'phone', 'cohort']
    readonly_fields = ['created_at', 'updated_at']
    actions = ['export_to_excel', 'mark_as_approved', 'mark_as_rejected']

    @admin.action(description='Xuất file Excel cho các đơn được chọn')
    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DS_Dang_Ky_ICPC"

        # Headers
        headers = [
            'STT',
            'Tài khoản HNUEOJ',
            'Họ và tên',
            'Mã sinh viên',
            'Lớp',
            'Khóa',
            'Email',
            'Số điện thoại (Zalo)',
            'Link Facebook',
            'Trình độ lập trình',
            'Thành tích / Kinh nghiệm',
            'Lý do & Mục tiêu',
            'Trạng thái',
            'Ghi chú Admin',
            'Thời gian đăng ký',
        ]
        ws.append(headers)

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 28

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        for idx, item in enumerate(queryset, start=1):
            username = item.user.username if item.user else 'Chưa liên kết'
            created_str = item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else ''
            
            row = [
                idx,
                username,
                item.full_name,
                item.student_id,
                item.academic_class,
                item.cohort,
                item.email,
                item.phone,
                item.facebook_url or '',
                item.get_programming_experience_display(),
                item.achievements or '',
                item.reason or '',
                item.get_status_display(),
                item.admin_note or '',
                created_str,
            ]
            ws.append(row)
            row_idx = idx + 1
            ws.row_dimensions[row_idx].height = 22

            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                if col_num in [1, 4, 6, 8, 13, 15]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Danh_Sach_Dang_Ky_ICPC.xlsx"'
        wb.save(response)
        return response

    @admin.action(description='Duyệt các đơn được chọn')
    def mark_as_approved(self, request, queryset):
        count = queryset.update(status=ICPCRegistration.STATUS_APPROVED)
        self.message_user(request, f"Đã duyệt thành công {count} đơn đăng ký.")

    @admin.action(description='Từ chối các đơn được chọn')
    def mark_as_rejected(self, request, queryset):
        count = queryset.update(status=ICPCRegistration.STATUS_REJECTED)
        self.message_user(request, f"Đã từ chối {count} đơn đăng ký.")
