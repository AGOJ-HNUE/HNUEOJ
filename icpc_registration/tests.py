from django.contrib.auth import get_user_model
from django.test import TestCase, Client
from django.urls import reverse
import openpyxl
from io import BytesIO

from .models import ICPCRegistration
from judge.models import Profile

User = get_user_model()


class ICPCRegistrationTests(TestCase):
    def setUp(self):
        self.client = Client(SERVER_NAME='oj.hnue.info.vn')
        self.user = User.objects.create_user(
            username='student1',
            email='student1@hnue.edu.vn',
            password='password123',
            first_name='Nguyen',
            last_name='Van A',
        )
        Profile.objects.get_or_create(user=self.user)

        self.staff_user = User.objects.create_user(
            username='admin1',
            email='admin1@hnue.edu.vn',
            password='adminpassword',
            is_staff=True,
            is_superuser=True,
        )
        Profile.objects.get_or_create(user=self.staff_user)

    def test_icpc_index_view(self):
        url = reverse('icpc_registration:index')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'TIỂU BAN ICPC')
        self.assertContains(response, 'co_cau_icpc')
        self.assertContains(response, 'Training đợt 1')

    def test_icpc_register_view_get(self):
        url = reverse('icpc_registration:index')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Form Đăng ký thành viên Tiểu ban ICPC')

    def test_icpc_register_view_post_success(self):
        url = reverse('icpc_registration:index')
        self.client.login(username='student1', password='password123')
        data = {
            'full_name': 'Tran Van B',
            'student_id': '725105099',
            'academic_class': 'K72A Sư phạm Tin',
            'cohort': 'K72',
            'email': 'stu725105099@hnue.edu.vn',
            'phone': '0987654321',
            'facebook_url': 'https://facebook.com/tranvanb',
            'programming_experience': ICPCRegistration.EXPERIENCE_INTERMEDIATE,
            'achievements': 'Giải nhì HSG Tin học cấp tỉnh',
            'reason': 'Muốn học hỏi thuật toán và đại diện trường thi ICPC',
        }
        response = self.client.post(url, data, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(ICPCRegistration.objects.filter(student_id='725105099').exists())
        reg = ICPCRegistration.objects.get(student_id='725105099')
        self.assertEqual(reg.full_name, 'Tran Van B')
        self.assertEqual(reg.status, ICPCRegistration.STATUS_PENDING)

    def test_duplicate_student_id_validation(self):
        self.client.login(username='student1', password='password123')
        ICPCRegistration.objects.create(
            full_name='Test Existing',
            student_id='725105001',
            academic_class='K71B',
            cohort='K71',
            email='stu725105001@hnue.edu.vn',
            phone='0912345678',
            reason='Sample reason',
            status=ICPCRegistration.STATUS_PENDING,
        )

        url = reverse('icpc_registration:index')
        data = {
            'full_name': 'Test Duplicate',
            'student_id': '725105001',
            'academic_class': 'K72A',
            'cohort': 'K72',
            'email': 'stu725105002@hnue.edu.vn',
            'phone': '0988888888',
            'reason': 'Sample reason 2',
        }
        response = self.client.post(url, data, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'đã có đơn đăng ký')

    def test_hnue_email_validation_and_duplicate(self):
        self.client.login(username='student1', password='password123')
        ICPCRegistration.objects.create(
            full_name='Student X',
            student_id='725105111',
            academic_class='K72A',
            cohort='K72',
            email='stu725105111@hnue.edu.vn',
            phone='0912345678',
            reason='Sample reason',
            status=ICPCRegistration.STATUS_PENDING,
        )

        url = reverse('icpc_registration:index')
        # 1. Non-HNUE email should fail
        bad_data = {
            'full_name': 'Student Y',
            'student_id': '725105222',
            'academic_class': 'K72A',
            'cohort': 'K72',
            'email': 'invalid_email@gmail.com',
            'phone': '0988888888',
            'reason': 'Sample reason',
        }
        resp1 = self.client.post(url, bad_data)
        self.assertEqual(resp1.status_code, 200)
        self.assertContains(resp1, 'Email sinh viên phải là email HNUE')

        # 2. Duplicate HNUE email should fail
        dup_data = {
            'full_name': 'Student Z',
            'student_id': '725105333',
            'academic_class': 'K72A',
            'cohort': 'K72',
            'email': 'stu725105111@hnue.edu.vn',
            'phone': '0988888888',
            'reason': 'Sample reason',
        }
        resp2 = self.client.post(url, dup_data)
        self.assertEqual(resp2.status_code, 200)
        self.assertContains(resp2, 'đã được sử dụng để đăng ký')

    def test_unauthenticated_user_cannot_register(self):
        self.client.logout()
        url = reverse('icpc_registration:index')
        data = {
            'full_name': 'Guest User',
            'student_id': '725105999',
            'academic_class': 'K73A',
            'cohort': 'K73',
            'email': 'guest@gmail.com',
            'phone': '0977777777',
            'reason': 'Guest trying to register',
        }
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/user/login', response.url)

    def test_admin_stats_access_control(self):
        url = reverse('icpc_registration:stats')
        # Non-logged in or non-staff should fail or redirect
        response = self.client.get(url)
        self.assertIn(response.status_code, [302, 403])

        # Staff user should succeed
        self.client.login(username='admin1', password='adminpassword')
        response_staff = self.client.get(url)
        self.assertEqual(response_staff.status_code, 200)
        self.assertContains(response_staff, 'TRANG QUẢN TRỊ VIÊN')

    def test_export_excel_view(self):
        ICPCRegistration.objects.create(
            full_name='Nguyen Van C',
            student_id='725105088',
            academic_class='K73A',
            cohort='K73',
            email='vanc@gmail.com',
            phone='0933333333',
            reason='Mục tiêu học tập',
            status=ICPCRegistration.STATUS_APPROVED,
        )

        self.client.login(username='admin1', password='adminpassword')
        url = reverse('icpc_registration:export_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = openpyxl.load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "DS_Dang_Ky_ICPC")
        # Check title in row 1 and headers in row 3
        self.assertIn('DANH SÁCH SINH VIÊN ĐĂNG KÝ', str(ws['A1'].value))
        self.assertEqual(ws.cell(row=3, column=3).value, 'Họ và tên')
        self.assertEqual(ws.cell(row=4, column=3).value, 'Nguyen Van C')

    def test_update_status_ajax(self):
        reg = ICPCRegistration.objects.create(
            full_name='Nguyen Van D',
            student_id='725105077',
            academic_class='K72B',
            cohort='K72',
            email='vand@gmail.com',
            phone='0944444444',
            reason='Lý do tham gia',
            status=ICPCRegistration.STATUS_PENDING,
        )

        self.client.login(username='admin1', password='adminpassword')
        url = reverse('icpc_registration:update_status', kwargs={'pk': reg.pk})
        response = self.client.post(url, {'status': 'approved'})
        self.assertEqual(response.status_code, 200)
        json_data = response.json()
        self.assertTrue(json_data['success'])
        
        reg.refresh_from_db()
        self.assertEqual(reg.status, ICPCRegistration.STATUS_APPROVED)
