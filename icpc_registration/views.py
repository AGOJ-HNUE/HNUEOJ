from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import ICPCRegistrationForm
from .models import ICPCRegistration


def is_staff_user(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def icpc_index(request):
    """
    Trang gộp duy nhất của Tiểu ban ICPC:
    Bao gồm Giới thiệu, Sơ đồ cơ cấu ảnh đính kèm, Kế hoạch hoạt động Vertical Timeline 
    và Form điền Đăng ký thành viên trực tiếp bên dưới.
    """
    initial_data = {}
    if request.user.is_authenticated:
        full_name = f"{request.user.first_name} {request.user.last_name}".strip()
        if full_name:
            initial_data['full_name'] = full_name
        else:
            initial_data['full_name'] = request.user.username
        
        if request.user.email:
            initial_data['email'] = request.user.email

    user_existing_reg = None
    if request.user.is_authenticated:
        user_existing_reg = ICPCRegistration.objects.filter(user=request.user).first()

    if request.method == 'POST':
        # Yêu cầu người dùng phải đăng nhập tài khoản mới được nộp đơn
        if not request.user.is_authenticated:
            messages.warning(
                request,
                'Bạn cần phải đăng nhập tài khoản HNUEOJ trước khi gửi đơn đăng ký!'
            )
            return redirect('/user/login?next=/clb-nvsp/icpc/#dang-ky')

        # Chặn nếu người dùng đã có đơn đăng ký liên kết với tài khoản này
        if user_existing_reg:
            messages.warning(
                request,
                'Tài khoản của bạn đã gửi đơn đăng ký rồi! Không thể gửi thêm đơn mới.'
            )
            return redirect('icpc_registration:index')

        form = ICPCRegistrationForm(request.POST)
        if form.is_valid():
            registration = form.save(commit=False)
            if request.user.is_authenticated:
                registration.user = request.user
            registration.save()
            messages.success(
                request,
                'Chúc mừng! Đơn đăng ký tham gia Tiểu ban ICPC của bạn đã được gửi thành công. Ban quản lý sẽ liên hệ với bạn trong thời gian sớm nhất!'
            )
            return redirect('icpc_registration:index')
        else:
            messages.error(
                request,
                'Có lỗi xảy ra trong quá trình điền form. Vui lòng kiểm tra lại các trường thông tin bên dưới!'
            )
    else:
        form = ICPCRegistrationForm(initial=initial_data)

    context = {
        'title': 'Tiểu ban ICPC - CLB Nghiệp vụ sư phạm - Khoa CNTT',
        'form': form,
        'user_existing_reg': user_existing_reg,
        'active_tab': 'index',
    }
    return render(request, 'icpc_registration/index.html', context)


@user_passes_test(is_staff_user)
def icpc_admin_stats(request):
    """Trang thống kê & Quản lý đơn đăng ký dành riêng cho Quản trị viên."""
    queryset = ICPCRegistration.objects.all()

    # Bộ lọc & Tìm kiếm
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    cohort_filter = request.GET.get('cohort', '').strip()
    experience_filter = request.GET.get('experience', '').strip()

    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(academic_class__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if cohort_filter:
        queryset = queryset.filter(cohort__iexact=cohort_filter)

    if experience_filter:
        queryset = queryset.filter(programming_experience=experience_filter)

    # Thống kê tổng hợp
    all_registrations = ICPCRegistration.objects.all()
    total_count = all_registrations.count()
    pending_count = all_registrations.filter(status=ICPCRegistration.STATUS_PENDING).count()
    approved_count = all_registrations.filter(status=ICPCRegistration.STATUS_APPROVED).count()
    rejected_count = all_registrations.filter(status=ICPCRegistration.STATUS_REJECTED).count()

    # Calculate percentages for progress bars
    pending_percent = round((pending_count / total_count * 100), 1) if total_count > 0 else 0
    approved_percent = round((approved_count / total_count * 100), 1) if total_count > 0 else 0
    rejected_percent = round((rejected_count / total_count * 100), 1) if total_count > 0 else 0

    cohort_stats = (
        all_registrations.values('cohort')
        .annotate(count=Count('id'))
        .order_by('cohort')
    )

    experience_stats = (
        all_registrations.values('programming_experience')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    available_cohorts = (
        all_registrations.values_list('cohort', flat=True)
        .distinct()
        .order_by('cohort')
    )

    context = {
        'title': 'Thống kê Quản trị viên - Đăng ký ICPC',
        'registrations': queryset,
        'total_count': total_count,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'pending_percent': pending_percent,
        'approved_percent': approved_percent,
        'rejected_percent': rejected_percent,
        'cohort_stats': cohort_stats,
        'experience_stats': experience_stats,
        'available_cohorts': available_cohorts,
        'search_query': search_query,
        'status_filter': status_filter,
        'cohort_filter': cohort_filter,
        'experience_filter': experience_filter,
        'experience_choices': ICPCRegistration.EXPERIENCE_CHOICES,
        'status_choices': ICPCRegistration.STATUS_CHOICES,
        'active_tab': 'stats',
    }
    return render(request, 'icpc_registration/stats.html', context)


@user_passes_test(is_staff_user)
def icpc_export_excel(request):
    """Xuất danh sách đơn đăng ký thành file Excel (.xlsx)."""
    queryset = ICPCRegistration.objects.all()

    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    cohort_filter = request.GET.get('cohort', '').strip()
    experience_filter = request.GET.get('experience', '').strip()

    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(academic_class__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if cohort_filter:
        queryset = queryset.filter(cohort__iexact=cohort_filter)

    if experience_filter:
        queryset = queryset.filter(programming_experience=experience_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DS_Dang_Ky_ICPC"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(name='Calibri', size=14, bold=True, color='1E3A8A')
    title_align = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A1:O1')
    ws['A1'] = "DANH SÁCH SINH VIÊN ĐĂNG KÝ THÀNH VIÊN TIỂU BAN ICPC - CLB NGHIỆP VỤ SƯ PHẠM KHOA CNTT"
    ws['A1'].font = title_font
    ws['A1'].alignment = title_align
    ws.row_dimensions[1].height = 30

    headers = [
        'STT',
        'Tài khoản HNUEOJ',
        'Họ và tên',
        'Mã sinh viên',
        'Lớp',
        'Khóa',
        'Email',
        'Số điện thoại (Zalo)',
        'Link Facebook',
        'Trình độ lập trình',
        'Thành tích / Kinh nghiệm',
        'Lý do & Mục tiêu tham gia',
        'Trạng thái',
        'Ghi chú Admin',
        'Thời gian đăng ký',
    ]

    ws.append([])
    ws.append(headers)
    ws.row_dimensions[3].height = 28

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    fill_approved = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_pending = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    fill_rejected = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for idx, item in enumerate(queryset, start=1):
        username = item.user.username if item.user else 'Chưa liên kết'
        created_str = item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else ''

        row_data = [
            idx,
            username,
            item.full_name,
            item.student_id,
            item.academic_class,
            item.cohort,
            item.email,
            item.phone,
            item.facebook_url or '',
            item.get_programming_experience_display(),
            item.achievements or '',
            item.reason or '',
            item.get_status_display(),
            item.admin_note or '',
            created_str,
        ]
        ws.append(row_data)
        row_idx = idx + 3
        ws.row_dimensions[row_idx].height = 24

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num in [1, 4, 6, 8, 13, 15]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

            if col_num == 13:
                if item.status == ICPCRegistration.STATUS_APPROVED:
                    cell.fill = fill_approved
                elif item.status == ICPCRegistration.STATUS_PENDING:
                    cell.fill = fill_pending
                elif item.status == ICPCRegistration.STATUS_REJECTED:
                    cell.fill = fill_rejected

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 48)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Danh_Sach_Dang_Ky_ICPC_CLB_NVSP.xlsx"'
    wb.save(response)
    return response


@user_passes_test(is_staff_user)
@require_POST
def icpc_update_status(request, pk):
    """Cập nhật trạng thái duyệt đơn qua AJAX."""
    registration = get_object_or_404(ICPCRegistration, pk=pk)
    new_status = request.POST.get('status', '').strip()
    admin_note = request.POST.get('admin_note', '').strip()

    valid_statuses = [choice[0] for choice in ICPCRegistration.STATUS_CHOICES]
    if new_status in valid_statuses:
        registration.status = new_status
        if admin_note:
            registration.admin_note = admin_note
        registration.save()
        return JsonResponse({
            'success': True,
            'status_display': registration.get_status_display(),
            'message': f'Đã cập nhật trạng thái đơn của {registration.full_name} thành {registration.get_status_display()}',
        })

    return JsonResponse({'success': False, 'error': 'Trạng thái không hợp lệ'}, status=400)
